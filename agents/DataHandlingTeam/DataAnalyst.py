"""
Data Analyst
Role: Performs data profiling, summarization, and column categorization to ensure structured, ready-to-model datasets.
Responsibilities:
    - Generate descriptive summaries of the dataset (shape, products, dates, missing values).
    - Collect and validate column configuration from the user.
    - Categorize columns using LLM-based classification and human-in-the-loop approval.
    - Identify distinct products and key dimensions for modeling readiness.
"""
from langchain_core.callbacks import streaming_stdout
from langgraph.graph import START, END, StateGraph
from langchain_ollama import ChatOllama
from itertools import zip_longest

from langgraph.types import Command
from agent_patterns.states import DataAnalystState, Feedback
from agent_patterns.structured_response import ColumnCategoriesResponse
from utils import utility
from utils.memory_handler import DataStore
from utils import theme_utility
from utils import chat_utility
import pandas as pd
from rich import print
from utils.theme_utility import console, log

import win32com.client as win32
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import shutil

DataValidationPrompt = utility.load_prompt_config(
    r"prompts\AgentPrompts.yaml",
    "DataValidationPrompt",
)


class DataAnalystAgent:
    def __init__(
        self,
        agent_name: str,
        agent_description: str,
        model: str,
        log_path: str = "logs/data_analyst.log",
    ):
        self.agent_name = agent_name
        self.agent_description = agent_description

        self.model = model
        self.llm = ChatOllama(model=self.model)
        self.log_path = log_path
        theme_utility.setup_console_logging(log_path)
        self.graph = self._build_graph()

    def DataSummaryNode(self, state: DataAnalystState):
        log("[medium_purple3]LOG: Starting Data Summarizer...[/]")
        with console.status(f"[plum1] Data Summarizer Node setting up...[/]", spinner="dots"):
            col_config = [
                "date_col",
                "product_col",
                "price_col",
                "revenue_col",
                "units_sold_col",
                "oos_col",
            ]
        try:
            actual_cols = list(DataStore.get_df("master_data").columns)
            rows = list(zip_longest(col_config, actual_cols))
            required_columns_df = pd.DataFrame(
                rows, columns=["COLUMN_CONFIG", "COLUMN_NAME"]
            )
            file_path = r"user_inputs\column_config.xlsx"
            chat_utility.user_input_excel(required_columns_df, file_path=file_path)

            console.print(f"\n[light_sky_blue1]Columns detected: [/]")
            theme_utility.print_items_as_panels(actual_cols)

            column_config_df = pd.read_excel(file_path)
            DataStore.set_df("column_config_df", column_config_df)
            column_config = dict(
                zip(column_config_df["COLUMN_CONFIG"], column_config_df["COLUMN_NAME"])
            )

            date_col = column_config["date_col"]
            product_col = column_config["product_col"]
            price_col = column_config["price_col"]

            data_summary = self.full_dataframe_summary(
                DataStore.get_df("master_data"), date_col, product_col, price_col
            )

            # --- BASIC INFO ---
            console.print("\n[bold underline]Basic Info[/bold underline]")
            basic_info = [
                ("Shape", data_summary.get("Shape")),
                ("Products", data_summary.get("Unique Products")),
            ]
            theme_utility.print_rich_table(basic_info, headers=["Metric", "Value"])

            # --- DATE INFO ---
            date_info = data_summary.get("Date Info", {})
            if date_info:
                console.print("\n[bold underline]Date Info[/bold underline]")
                theme_utility.print_rich_table(
                    list(date_info.items()), headers=["Metric", "Value"]
                )

            # --- MISSING VALUES ---
            missing = data_summary.get("Missing Values", {})
            if missing:
                console.print("\n[bold underline]Missing Values[/bold underline]")
                rows = [(k, v["Count"], f"{v['Percent']:.2f}%") for k, v in missing.items()]
                theme_utility.print_rich_table(
                    rows, headers=["Column", "Missing Count", "Missing %"]
                )

            # --- UNIQUE & MOST FREQUENT ---
            unique_freq = data_summary.get("Unique and Most Frequent", {})
            if unique_freq:
                console.print("\n[bold underline]Unique and Most Frequent[/bold underline]")
                rows = [
                    (k, v["Unique Values"], v["Most Frequent"])
                    for k, v in unique_freq.items()
                ]
                theme_utility.print_rich_table(
                    rows, headers=["Column", "Unique Values", "Most Frequent"]
                )

            # --- PRODUCT SUMMARY ---
            prod_summary = data_summary.get("Product Summary", {})
            if prod_summary:
                console.print("\n[bold underline]Product Summary[/bold underline]")
                rows = [
                    (k, v["Start Date"], v["End Date"], v["Data Points"], v["Avg Price"])
                    for k, v in prod_summary.items()
                ]
                theme_utility.print_rich_table(
                    rows,
                    headers=[
                        "Product",
                        "Start Date",
                        "End Date",
                        "Data Points",
                        "Avg Price",
                    ],
                )
            utility.save_to_memory_file('data_summary.txt', str(data_summary))
            log(f"[medium_purple3]LOG: Saved data summary to memory[/]")
            log("[green3]LOG: Data generated report [/]")
            asst_message = chat_utility.build_message_structure(role = "assistant", message = "Data Summarization process complete")
            return Command(
                goto = "colCategorizeNode",
                update = {
                    "data_summary": self._cast_to_builtin_types(data_summary),
                    "messages": [asst_message]
                    }
            )
        except Exception as e:
            return Command(
                goto = END, 
                update = {
                    "messages": [chat_utility.build_message_structure(role = "assistant", message = f"Error while summarizing data :{e}")]
                }
            )

    def ColumnCatogerizerNode(self, state: DataAnalystState):
        log("[medium_purple3]LOG: Starting Column Catogorizing[/]")
        with console.status(
            f"[plum1] Column Catogorizer Node setting up...[/]", spinner="dots"
        ):
            llm_structured = self.llm.with_structured_output(ColumnCategoriesResponse)
            messages = chat_utility.build_message_structure(
                role="system", message=DataValidationPrompt["ColumnCatogerizer"]
            )
            messages = [messages] + [
                chat_utility.build_message_structure(
                    role="user", message=f"LIST OF COLUMNS: {list(DataStore.get_df("master_data").columns)}"
                )
            ]
            try:
                if state["user_feedback"].category == "retry":
                    chat_utility.append_to_structure(
                        history=messages,
                        role="assistant",
                        message=f"THOUGHT : {state['user_feedback'].thought}",
                    )
                    chat_utility.append_to_structure(
                        history=messages,
                        role="user",
                        message=f"FEEDBACK: {state['user_feedback'].feedback}",
                    )
            except:
                pass
            response = llm_structured.invoke(messages)
            theme_utility.display_response(response.thought_process, title="Thought",border_style='light_steel_blue')
            theme_utility.display_response(
                ", ".join(list(DataStore.get_df("master_data").columns)), title="ALL COLUMNS", border_style='light_yellow3'
            )
        log("[green3]LOG: Column Catogories Generated[/]")
        asst_message = chat_utility.build_message_structure(role = "assistant", message = "Column Catogories Generated")
        return {
            "column_categories": self._cast_to_builtin_types(dict(response.column_categories)),
            "messages": [asst_message]
            }

    def colCatApprovalNode(self, state: DataAnalystState):
        log("[medium_purple3]LOG: Column Category Approval required[/]")
        with console.status(
            f"[plum1] Column Category Approval Node setting up...[/]", spinner="dots"
        ):
            console.print(f"\n[light_sky_blue1]Printing column categories:[/]\n")
            theme_utility.print_dictionary(
                state["column_categories"], title="column_categories"
            )
        approved, suggestion = chat_utility.ask_user_approval(
            agent_name="Category Node"
        )
        if approved is True:
            user_feedback_llm = Feedback(category="approve", feedback="", thought="")
            log(f"[medium_purple3] Category Approval Node approved, printing data dictionary[/]")
        elif approved is False:
            user_feedback_llm = Feedback(category="retry", feedback="", thought="")
            log(f"[red3] Category Approval Node Denied, retrying...[/]")
        else:
            feedback_llm = self.llm.with_structured_output(Feedback)
            message = chat_utility.build_message_structure(
                role="system", message=DataValidationPrompt["ApprovalNode"]
            )
            message = [message] + [
                chat_utility.build_message_structure(role="user", message=suggestion)
            ]

            user_feedback_llm = feedback_llm.invoke(message)
            log(f"[medium_purple3] Suggestion received[/] : [turquoise4]{suggestion}[/]")

        return {"user_feedback": user_feedback_llm}

    def _colCatDecisionNode(self, state: DataAnalystState):
        if state["user_feedback"].category == "approve":
            utility.save_in_memory(
                "column_categories", to_save=state["column_categories"], desc = "Column categories"
            )
            return "approved"
        elif state["user_feedback"].category == "retry":
            return "retry"
        else:
            return "suggested_or_denied"

    def distinctProductNode(self, state: DataAnalystState):
        log("[medium_purple3]LOG: Identifying Distinct Products[/]")
        with console.status(f"[plum1] Distinct Product Identifier Node setting up...[/]", spinner="dots"):
            distinct_products = DataStore.get_df("master_data")[state["column_categories"]["product_col"]].dropna().unique().tolist()
        console.print(f"\n[light_sky_blue1]Distinct Products:[/]")
        theme_utility.print_items_as_panels(distinct_products)
        log("[green3]LOG: Distinct Product Identification completed[/]")
        asst_message = chat_utility.build_message_structure(role = "assistant", message = "Distinct Product Identification completed")
        state['messages'].append(asst_message)
        state['distinct_products'] = self._cast_to_builtin_types(distinct_products)
        DataStore.set_str("distinct_products",str(distinct_products))
        state['completed'] = True
        state['user_feedback'] = None
        return state


    def generate_excel(self, state: DataAnalystState):
        log("[medium_purple3]LOG: Generate Excel for interactive Dashboard[/]")
        with console.status(f"[plum1] Generating interactive dashboard...[/]", spinner="dots"):
            column_config_df = pd.read_excel(r"C:\Users\nigam\Documents\AutoMMM\user_inputs\column_config.xlsx")
            column_config = dict(zip(column_config_df["COLUMN_CONFIG"], column_config_df["COLUMN_NAME"]))
            date_col = column_config["date_col"]
            product_col = column_config["product_col"]

            df = DataStore.get_df("master_data")
            csv_path = "output/temp_master_data.csv"
            os.makedirs("output", exist_ok=True)
            df.to_csv(csv_path, index=False)
            base_file_path = r"C:\Users\nigam\Documents\AutoMMM\utils\BASE.xlsx"
            destination = r"C:\Users\nigam\Documents\AutoMMM\output\analysis.xlsx"
            shutil.copy(base_file_path, destination)
            df["KEY"] = (df[product_col].astype(str) + (df[date_col] - pd.Timestamp("1899-12-30")).dt.days.astype(str))
            cols = ["KEY"] + [col for col in df.columns if col != "KEY"]
            df = df[cols]
        
        wb_path = r"C:\Users\nigam\Documents\AutoMMM\output\analysis.xlsx"

        sheet_name = "raw_data"
        with pd.ExcelWriter(wb_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        xlApp = win32.Dispatch('Excel.Application')
        xlApp.Visible = True
        # Write lookup formulas
        rows, cols = df.shape
        last_letter = get_column_letter(cols)

        wb = xlApp.Workbooks.Open(wb_path)
        # Backend Sheet
        try:
            ws_to_delete = wb.Sheets('backend')
            ws_to_delete.Delete()
        except:
            pass
        wb.Sheets.Add().Name = 'backend'
        ws = wb.Sheets('backend')

        header_cell = ws.Range("C5")
        header_cell.Value = "UNIQUE_PRODUCTS"
        header_cell.Font.Bold = True 
        unique_products = df[product_col].unique()
        start_cell = ws.Range("C6")
        for i, val in enumerate(unique_products):
            start_cell.Offset(i+1, 1).Value = val 

        header_cell_columns = ws.Range("E5")
        header_cell_columns.Value = "COLUMNS"
        header_cell_columns.Font.Bold = True
        columns = df.columns.tolist()
        start_cell_columns = ws.Range("E6")
        for i, col_name in enumerate(columns):
            start_cell_columns.Offset(i+1, 1).Value = col_name
        ws.Columns.AutoFit()

        # TREND BASE
        try:
            ws_to_delete = wb.Sheets('trend_base')
            ws_to_delete.Delete()
        except:
            pass

        wb.Sheets.Add().Name = 'trend_base'
        ws = wb.Sheets('trend_base')

        ws.Rows(1).Font.Bold = True
        copy_data_formula = [
            [f"=raw_data!{get_column_letter(col)}{row}" for col in range(1, cols + 1)]
            for row in range(1, rows + 1)
        ]

        ws.Range(ws.Cells(1, 1), ws.Cells(rows, cols)).Formula = copy_data_formula
        ws.Columns.AutoFit()

        # chart Sheet
        ws = wb.Sheets("analysis")

        ws.Range("V3").Formula = "=MINIFS(trend_base!B:B, trend_base!C:C, analysis!$C$6)"
        ws.Range("V4").Formula = "=MAXIFS(trend_base!B:B, trend_base!C:C, analysis!$C$6)"

        n_rows = (df[date_col].max() - df[date_col].min()).days // 7 + 1
        start_row = 7
        end_row = start_row + n_rows - 1

        cols_V = [["=$C$6"] for _ in range(n_rows)]
        cols_W = [["=$V$3 + 7*(ROW()-7)"] for _ in range(n_rows)]  # Simple relative formula
        cols_U = [[f"=V{start_row + i}&W{start_row + i}"] for i in range(n_rows)]
        cols_X = [[
            f"=XLOOKUP(U{start_row + i},trend_base!$A$2:$A$313,"
            f"XLOOKUP($X$6,trend_base!$D$1:$P$1,trend_base!$D$2:$P$313))"
        ] for i in range(n_rows)]
        cols_Y = [[
            f"=XLOOKUP(U{start_row + i},trend_base!$A$2:$A$313,"
            f"XLOOKUP($Y$6,trend_base!$D$1:$P$1,trend_base!$D$2:$P$313))"
        ] for i in range(n_rows)]

        ws.Range(f"U{start_row}:U{end_row}").Formula = cols_U
        ws.Range(f"V{start_row}:V{end_row}").Formula = cols_V
        ws.Range(f"W{start_row}:W{end_row}").Formula = cols_W
        ws.Range(f"X{start_row}:X{end_row}").Formula = cols_X
        ws.Range(f"Y{start_row}:Y{end_row}").Formula = cols_Y

        if os.path.exists(csv_path):
            os.remove(csv_path)

        wb.Save()
        return state


    def _build_graph(self):
        g = StateGraph(DataAnalystState)
        g.add_node("DataSummaryNode", self.DataSummaryNode)
        g.add_node("colCategorizeNode", self.ColumnCatogerizerNode)
        g.add_node("colCatApprovalNode", self.colCatApprovalNode)
        g.add_node("distinctProductNode", self.distinctProductNode)
        g.add_node("generate_excel", self.generate_excel)

        g.add_edge(START, "DataSummaryNode")
        # g.add_edge("DataSummaryNode", "colCategorizeNode")
        g.add_edge("colCategorizeNode", "colCatApprovalNode")
        g.add_conditional_edges(
            "colCatApprovalNode",
            self._colCatDecisionNode,
            {
                "approved": "distinctProductNode",
                "retry": "colCategorizeNode",
                "suggested_or_denied": "colCatApprovalNode",
            },
        )
        g.add_edge("distinctProductNode", "generate_excel")
        g.add_edge("generate_excel", END)
        
        return g.compile(name=self.agent_name)

    def full_dataframe_summary(self, df, date_col, product_col, price_col):
        summary = {}

        summary["Shape"] = df.shape
        summary["Unique Products"] = df[product_col].nunique()

        missing = df.isnull().sum()
        missing_percent = (missing / len(df)) * 100
        summary["Missing Values"] = pd.DataFrame(
            {"Count": missing, "Percent": missing_percent}
        ).to_dict(orient="index")

        unique_info = {}
        for col in df.columns:
            unique_vals = df[col].nunique(dropna=False)
            top_value = df[col].mode(dropna=False)
            unique_info[col] = {
                "Unique Values": int(unique_vals),
                "Most Frequent": top_value.iloc[0] if not top_value.empty else None,
            }
        summary["Unique and Most Frequent"] = unique_info

        if date_col in df.columns:
            try:
                df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
                date_series_df = df[df[product_col] == df[product_col].unique()[0]]

                date_series = date_series_df[date_col].dropna().sort_values()
                if not date_series.empty:
                    inferred_freq = pd.infer_freq(date_series)
                    summary["Date Info"] = {
                        "Date Column": date_col,
                        "Min Date": str(date_series.min()),
                        "Max Date": str(date_series.max()),
                        "Inferred Frequency": inferred_freq or "Irregular",
                    }
            except Exception as e:
                summary["Date Info"] = {"Error": str(e)}

        product_summary = {}
        for product, group in df.groupby(product_col):
            group_dates = pd.to_datetime(group[date_col], errors="coerce").dropna()
            price_mean = group[price_col].mean() if price_col in group.columns else None
            product_summary[str(product)] = {
                "Start Date": str(group_dates.min()) if not group_dates.empty else None,
                "End Date": str(group_dates.max()) if not group_dates.empty else None,
                "Data Points": len(group),
                "Avg Price": float(price_mean) if price_mean is not None else None,
            }
        summary["Product Summary"] = product_summary

        return summary
    
    def _cast_to_builtin_types(self, obj):
        import numpy as np
        import pandas as pd

        if isinstance(obj, dict):
            return {str(k): self._cast_to_builtin_types(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self._cast_to_builtin_types(i) for i in obj]
        elif isinstance(obj, (np.integer, np.int64, np.int32)):
            return int(obj)
        elif isinstance(obj, (np.floating, np.float64, np.float32)):
            return float(obj)
        elif isinstance(obj, (np.bool_, bool)):
            return bool(obj)
        elif isinstance(obj, (np.datetime64, pd.Timestamp)):
            return str(obj)
        elif isinstance(obj, pd.DataFrame):
            return obj.to_dict(orient="records")
        elif isinstance(obj, pd.Series):
            return obj.tolist()
        else:
            return obj
