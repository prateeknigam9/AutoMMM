import win32com.client as win32
import xlsxwriter
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pandas as pd
import sys
import os
import shutil

column_config_df = pd.read_excel(r"C:\Users\nigam\Documents\AutoMMM\user_inputs\column_config.xlsx")
column_config = dict(zip(column_config_df["COLUMN_CONFIG"], column_config_df["COLUMN_NAME"]))
date_col = column_config["date_col"]
product_col = column_config["product_col"]


df = pd.read_excel(r"C:\Users\nigam\Documents\AutoMMM\data_to_model.xlsx", sheet_name = "Sheet1")


df = DataStore.get_df("master_data")
csv_path = "output/temp_master_data.csv"
os.makedirs("output", exist_ok=True)
df.to_csv(csv_path, index=False)

base_file_path = r"C:\Users\nigam\Documents\AutoMMM\utils\BASE.xlsx"
destination = r"C:\Users\nigam\Documents\AutoMMM\output\analysis.xlsx"


shutil.copy(base_file_path, destination)

df["KEY"] = (df[product_col].astype(str) + (df[date_col] - pd.Timestamp("1899-12-30")).dt.days.astype(str))
cols = ["KEY"] + [col for col in df.columns if col != "KEY"]
df = df[cols]


wb_path = r"C:\Users\nigam\Documents\AutoMMM\output\analysis.xlsx"

sheet_name = "raw_data"
with pd.ExcelWriter(wb_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, index=False, sheet_name=sheet_name)

xlApp = win32.Dispatch('Excel.Application')
xlApp.Visible = True

# Write lookup formulas
rows, cols = df.shape
last_letter = get_column_letter(cols)


wb = xlApp.Workbooks.Open(wb_path)
# Backend Sheet
try:
    ws_to_delete = wb.Sheets('backend')
    ws_to_delete.Delete()
except:
    pass

wb.Sheets.Add().Name = 'backend'
ws = wb.Sheets('backend')

header_cell = ws.Range("C5")
header_cell.Value = "UNIQUE_PRODUCTS"
header_cell.Font.Bold = True 
unique_products = df[product_col].unique()
start_cell = ws.Range("C6")
for i, val in enumerate(unique_products):
    start_cell.Offset(i+1, 1).Value = val 

header_cell_columns = ws.Range("E5")
header_cell_columns.Value = "COLUMNS"
header_cell_columns.Font.Bold = True
columns = df.columns.tolist()
start_cell_columns = ws.Range("E6")
for i, col_name in enumerate(columns):
    start_cell_columns.Offset(i+1, 1).Value = col_name
ws.Columns.AutoFit()

# TREND BASE
try:
    ws_to_delete = wb.Sheets('trend_base')
    ws_to_delete.Delete()
except:
    pass

wb.Sheets.Add().Name = 'trend_base'
ws = wb.Sheets('trend_base')

ws.Rows(1).Font.Bold = True
copy_data_formula = [
    [f"=raw_data!{get_column_letter(col)}{row}" for col in range(1, cols + 1)]
    for row in range(1, rows + 1)
]

ws.Range(ws.Cells(1, 1), ws.Cells(rows, cols)).Formula = copy_data_formula
ws.Columns.AutoFit()


# chart Sheet
ws = wb.Sheets("analysis")

ws.Range("V3").Formula = "=MINIFS(trend_base!B:B, trend_base!C:C, analysis!$C$6)"
ws.Range("V4").Formula = "=MAXIFS(trend_base!B:B, trend_base!C:C, analysis!$C$6)"

n_rows = (df[date_col].max() - df[date_col].min()).days // 7 + 1
start_row = 7
end_row = start_row + n_rows - 1

cols_V = [["=$C$6"] for _ in range(n_rows)]
cols_W = [["=$V$3 + 7*(ROW()-7)"] for _ in range(n_rows)]  # Simple relative formula
cols_U = [[f"=V{start_row + i}&W{start_row + i}"] for i in range(n_rows)]
cols_X = [[
    f"=XLOOKUP(U{start_row + i},trend_base!$A$2:$A$313,"
    f"XLOOKUP($X$6,trend_base!$D$1:$P$1,trend_base!$D$2:$P$313))"
] for i in range(n_rows)]
cols_Y = [[
    f"=XLOOKUP(U{start_row + i},trend_base!$A$2:$A$313,"
    f"XLOOKUP($Y$6,trend_base!$D$1:$P$1,trend_base!$D$2:$P$313))"
] for i in range(n_rows)]


ws.Range(f"U{start_row}:U{end_row}").Formula = cols_U
ws.Range(f"V{start_row}:V{end_row}").Formula = cols_V
ws.Range(f"W{start_row}:W{end_row}").Formula = cols_W
ws.Range(f"X{start_row}:X{end_row}").Formula = cols_X
ws.Range(f"Y{start_row}:Y{end_row}").Formula = cols_Y

if os.path.exists(csv_path):
    os.remove(csv_path)