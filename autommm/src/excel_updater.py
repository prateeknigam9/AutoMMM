import win32com.client as win32
import xlsxwriter
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pandas as pd
import sys
import os

# ---- TEMP FIX FOR MODULE IMPORTS ----
# Adds project root (3 levels up from this file) to sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "../../..")))


# https://learndataanalysis.org/automate-excel-pivot-table-with-python/
# https://xlsxwriter.readthedocs.io/contents.html

def clear_pts(ws):
    for pt in ws.PivotTables():
        pt.TableRange2.Clear()

def insert_pt_fields(pt, row_fields, col_fields, data_fields):
    field_rows = {}
    field_cols = {}
    field_values = {}

    for idx, fld in enumerate(row_fields, start=1):
        field_rows[fld] = pt.PivotFields(fld)
        field_rows[fld].Orientation = 1  # xlRowField
        field_rows[fld].Position = idx
        field_rows[fld].Subtotals = tuple(False for _ in range(12))

    for idx, fld in enumerate(col_fields, start=1):
        field_cols[fld] = pt.PivotFields(fld)
        field_cols[fld].Orientation = 2  # xlColumnField
        field_cols[fld].Position = idx
        field_cols[fld].Subtotals = tuple(False for _ in range(12))

    for src, field_name, func, numfmt in data_fields:
        field_values[field_name] = pt.PivotFields(src)
        field_values[field_name].Orientation = 4  # xlDataField
        field_values[field_name].Function = func
        field_values[field_name].NumberFormat = numfmt

def create_pivot_table(data_sheet, report_sheet, report_sheet_ref:str, title: str, 
                       row_fields: list, col_fields: list, data_fields: list[tuple],
                       grand_total: bool = False, row_total: bool = False,
                       report_axis_layout: int = 1, 
                       table_style: str = "PivotStyleMedium9", wb=None
                       ):

    clear_pts(report_sheet)

    pt_cache = wb.PivotCaches().Create(1, data_sheet.Range("A1").CurrentRegion)
    pt = pt_cache.CreatePivotTable(report_sheet.Range(report_sheet_ref), title)

    pt.ColumnGrand = grand_total
    pt.RowGrand = row_total
    pt.RowAxisLayout(report_axis_layout)
    pt.TableStyle2 = table_style

    insert_pt_fields(pt, row_fields, col_fields, data_fields)

    pt.RepeatAllLabels(2)  # Excel constant = 2 :contentReference[oaicite:2]{index=2}

    return pt

def main():
    xlApp = win32.Dispatch('Excel.Application')
    xlApp.Visible = True

    wb_path = r"C:\Users\nigam\OneDrive\Documents\university_classes\AutoMMM\autommm\data\automated.xlsx"
    wb = xlApp.Workbooks.Open(wb_path)

    ws_data = wb.Worksheets("data")
    ws_report = wb.Worksheets("report")

    create_pivot_table(
        data_sheet=ws_data,
        report_sheet=ws_report,
        report_sheet_ref="B3",
        title="myreportsummary",
        row_fields=['date','sku'],
        col_fields=[],
        data_fields=[
            ("sales", "Total Sales", -4157, "$#,##0"),
            ("units", "Units Count", -4112, "#,##0")
        ],
        grand_total=False,
        row_total=False,
        report_axis_layout=1,
        table_style="PivotStyleMedium9",
        wb=wb
    )

    # Save workbook if changes were made
    # wb.Save()
    # # Close workbook and quit Excel
    # wb.Close(SaveChanges=False)
    # xlApp.Quit()


# def master_data_sheet(wb_path : str, df:pd.DataFrame):
    
#     with pd.ExcelWriter(wb_path, engine='xlsxwriter') as writer:
#         df.to_excel(writer, sheet_name='raw_data', index=False)
#         workbook = writer.book
# #         worksheet = workbook.add_worksheet('master_data')

#         data_sheet_name = 'raw_data'
#         rows, cols = df.shape
#         last_letter = get_column_letter(cols)

#         # Format entire column B - YYYY-MM-DD
#         worksheet.set_column('B:B', None, workbook.add_format({'num_format': 'yyyy-mm-dd'}))

#         # KEY col
#         worksheet.write('A1',"KEY")
#         for row in range(1, rows + 1):  # Skip header
#             formula = f'=C{row + 1}&B{row + 1}'
#             worksheet.write_formula(row, 0, formula)
        
#         worksheet.write('B1', f"={data_sheet_name}!A1:{last_letter}{rows + 1}")

def master_data_sheet(wb_path: str, df: pd.DataFrame):
    # Load the workbook
    wb = load_workbook(wb_path)
    if 'master_data' in wb.sheetnames:
        del wb['master_data']

    # Create a new worksheet
    ws = wb.create_sheet('master_data')

    data_sheet_name = 'raw_data'
    rows, cols = df.shape
    last_letter = get_column_letter(cols)

    for row in range(2, 1001):  # Skipping header in B1
        cell = ws[f'B{row}']
        cell.number_format = 'yyyy-mm-dd'


    # Write static values and headers
    ws['A1'] = "KEY"
    for row in range(2, rows + 2):  # 2-based Excel row index
        ws[f'A{row}'] = f'=C{row}&B{row}'
    
    ws['B1'] = f"={data_sheet_name}!A1:{last_letter}{rows + 1}"

    wb.save(wb_path)

def trend_base_sheet(wb_path: str, df: pd.DataFrame):
    # Load the workbook
    wb = load_workbook(wb_path)
    if 'trend_base' in wb.sheetnames:
        del wb['trend_base']

    # Create a new worksheet
    ws = wb.create_sheet('trend_base')

    # Write static values and headers
    ws['B1'] = "='trend analysis'!C3"

    ws['A3'] = "KEY"
    ws['B3'] = "product_id"
    ws['C3'] = "weekend"
    ws['D3'] = "units_sold"

    ws['E2'] = "KPI - 1"
    ws['F2'] = "KPI - 2"
    ws['E3'] = "='trend analysis'!$C$5"
    ws['F3'] = "='trend analysis'!$C$6"

    # Write formulas starting from row 4
    ws['C4'] = "=SEQUENCE((metadata!$C$4-metadata!$C$3)/7,1,metadata!$C$3,7)"
    ws['B4'] = "=$B$1"
    ws['A4'] = "=B4&C4"

    for row in range(4, 1001):  # Skipping header in B1
        cell = ws[f'B{row}']
        cell.number_format = 'yyyy-mm-dd'

    # Write lookup formulas
    rows, cols = df.shape
    last_letter = get_column_letter(cols)

    ws['D4'] = f"=XLOOKUP($A4,master_data!$A$2:$A${rows + 1},XLOOKUP(D$3,master_data!$B$1:${last_letter}$1,master_data!$B$2:${last_letter}${rows + 1}))"
    ws['E4'] = f"=XLOOKUP($A4,master_data!$A$2:$A${rows + 1},XLOOKUP(E$3,master_data!$B$1:${last_letter}$1,master_data!$B$2:${last_letter}${rows + 1}))"
    ws['F4'] = f"=XLOOKUP($A4,master_data!$A$2:$A${rows + 1},XLOOKUP(F$3,master_data!$B$1:${last_letter}$1,master_data!$B$2:${last_letter}${rows + 1}))"

    # Save the workbook
    wb.save(wb_path)


def format():
    wb_path = r"C:\Users\nigam\OneDrive\Documents\university_classes\AutoMMM\autommm\data\try.xlsx"
    df = pd.read_excel(r"C:\Users\nigam\OneDrive\Documents\university_classes\AutoMMM\autommm\data\data_analysis.xlsx", sheet_name="master_data")
    df = df[['date_week','product_id', 'year', 'month', 'dayofyear', 'branded_clicks',
       'branded_spends', 'branded_clicks_adstock',
       'branded_clicks_adstock_saturated', 'nonbranded_clicks',
       'nonbranded_spends', 'nonbranded_clicks_adstock',
       'nonbranded_clicks_adstock_saturated', 'oos', 'trend', 'cs', 'cc',
       'seasonality', 'event', 'intercept', 'epsilon', 'price',
       'log_price', 'insta_clicks', 'insta_spends', 'insta_clicks_adstock',
       'insta_clicks_adstock_saturated', 'fb_clicks', 'fb_spends',
       'fb_clicks_adstock', 'fb_clicks_adstock_saturated', 'units_sold',
       'revenue']]
    
    master_data_sheet(wb_path, df)
    trend_base_sheet(wb_path, df)
        # workbook.close()
    
if __name__ == "__main__":
    # main()
    format()
